from selenium import webdriver
from selenium.webdriver.common.by import By
from collections import OrderedDict
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import pandas as pd
import time
from datetime import date, datetime, timedelta
import re
import openpyxl
from openpyxl import workbook

# d1 = date.today()
# d = d1 - timedelta(days = 3)
# d2 = d1 - timedelta(days = 2)
# bin_day = d.strftime("%m/%d/%Y")
# bin_day2 = d2.strftime("%m/%d/%Y")
driver = webdriver.Chrome(executable_path=r"C:\Users\Public\AppData\chromedriver.exe")
driver.get("https://tt.amazon.com/search?category=&assigned_group=WW-Concession-Reduction&status=Sheet1&impact=&assigned_individual=&requester_login=&login_name=&cc_email=&phrase_search_text=&keyword_bq=&exact_bq=&or_bq1=&or_bq2=&or_bq3=&exclude_bq=&create_date=&modified_date=12%2F10%2F2020%2C12%2F11%2F2020&tags=&case_type=&building_id=&search=Search%21")
driver.maximize_window()
time.sleep(2)
# driver.find_element(By.LINK_TEXT, "Search").click()
# time.sleep(2)
# driver.find_element(By.XPATH,"""//*[@id="assigned_group_1"]""").click()
# driver.find_element(By.XPATH,"""//*[@id="assigned_group_1"]""").send_keys("WW-Concession-Reduction")
# time.sleep(2)
# driver.find_element(By.XPATH, """/html/body/div[1]/div[3]/form/dl[1]/div[1]/dd[3]/select""").click()
# time.sleep(2)
# driver.find_element(By.ID, "modified_date_from").send_keys(bin_day)
# driver.find_element(By.ID, "modified_date_to").send_keys(bin_day2)
# driver.find_element(By.ID, "dates-etc-ml-link").click()
# time.sleep(3)
# drop_reason = Select(driver.find_element(By.ID, "misc-field-dropdown"))
# drop_reason.select_by_visible_text("Pending reason")
# time.sleep(2)
# driver.find_element(By.ID, "pending_reason").send_keys("Verification of fix")
# time.sleep(2)
# driver.find_element(By.ID, "submit-button").click()
# time.sleep(2)
# table_Html  = driver.find_element(By.ID, "search_results").get_attribute("outerHTML")
# table = pd.read_html(table_Html)[0]
# Case_id = table[["Case ID","Asin"]]
# Case_id.to_excel(r'C:\Users\kamesk\Desktop\Bin\Bin_check_File.xlsx', index = False)


def getRowCount(file,sheetName):
	workbook = openpyxl.load_workbook(file)
	sheet = workbook.get_sheet_by_name(sheetName)
	return(sheet.max_row)

def getColumnCount (file,sheetName):
	workbook = openpyxl.load_workbook(file)
	sheet = workbook.get_sheet_by_name(sheetName)
	return(sheet.max_column)

def readData(file,sheetName,rownum,columnno):
	workbook = openpyxl.load_workbook(file)
	sheet = workbook.get_sheet_by_name(sheetName)
	return sheet.cell(row=rownum, column=columnno).value

def writeData(file,sheetName,rownum,columnno,data):
	workbook = openpyxl.load_workbook(file)
	sheet = workbook.get_sheet_by_name(sheetName)
	sheet.cell(row=rownum, column=columnno).value = data
	workbook.save(file)

# path3 = r"C:\Users\kamesk\Desktop\Bin\Bin_Master.xlsx"
path2 = r"C:\Users\kamesk\Desktop\Bin\Bin_check_File_FC.xlsx"
path = r"C:\Users\kamesk\Desktop\Bin\Bin_check_File.xlsx"

rows = getRowCount(path,'Sheet1')
rows2 = getRowCount(path2,'Sheet1')
# rows3 = getRowCount(path3,'Sheet1')


for r in range(2,rows+1):
    try:
        writeData(path, "Sheet1", r, 399, " ")
    except PermissionError:
        driver.execute_script("alert('PLEASE CLOSE THE INPUT EXCEL AND RUN AGAIN');")
        time.sleep(3)
    else:
        pass
    Ticket_num = readData(path, "Sheet1", r, 1)
    Asin_num = readData(path, "Sheet1", r, 2)
    if Ticket_num == None:
        driver.execute_script("alert('Process Completed');")
        break
    else:
        pass

    txt = driver.find_element(By.ID, "u_text_search").send_keys(Ticket_num)
    time.sleep(4)
    driver.find_element(By.CSS_SELECTOR, "#u_remedy_lookup > button").click()
    time.sleep(4)
    driver.find_element(By.LINK_TEXT, "WW-Concession-Reduction").click()
    time.sleep(4)
    driver.find_element(By.ID, "tab_correspondence").click()
    time.sleep(4)
    Content = driver.find_element(By.CLASS_NAME, "text_to_format").text
    page_1 = Content.replace("`", "")
    page_1 = page_1.replace("'", "")
    page_1 = page_1.replace("-", " ")
    page_1 = page_1.lower()
    line_sp = page_1
    # try:
    #     T_s = ["Necessary:"]
    #     for ttem in line_sp.split("\n\n"):
    #         for index in range(0, len(T_s)):
    #             if T_s[index] in ttem:
    #                 action_t_s = ttem
    #                 break
    #             else:
    #                 pass
    # except NoSuchElementException:
    #     pass
    # else:
    #     pass
    # tvv = re.split(':|\*|`|\*|\n', action_t_s)
    # tv = [i for item in tvv for i in item.split()]
    # finalt = []
    # for elementt in tv:
    #     finalt.append(elementt.strip())
    # finalt = [y for y in finalt if y]
    # primt = ['bin']
    # prit = all(itemt in finalt for itemt in primt)

    try:
        update = ['associates']
        for itemu in line_sp.split("\n\n"):
            for index in range(0, len(update)):
                if update[index] in itemu:
                    action_up = itemu
                    break
                else:
                    pass
    except:
        writeData(path, "Sheet1", r, 3 , "Blurb Incorrect")
        continue
    else:
        pass
    try:
        tuv = re.split(':|\*|`|\*|\n', action_up)
    except:
        writeData(path, "Sheet1", r, 3, "Blurb Incorrect")
        continue
    else:
        pass
    uv = [k for itemu in tuv for k in itemu.split()]
    final = []
    for element in uv:
        final.append(element.strip())
    final = [x for x in final if x]
    prim = ['bin']
    primv = ['vendor']
    pri = all(item in final for item in prim)
    priv = all(item in final for item in primv)

    if pri is True:
        driver.find_element(By.ID, "tab_related_items").click()
        time.sleep(3)
        try:
            Ri = driver.find_element(By.XPATH, """/html/body/div[1]/form/div[2]/div[2]/div[6]/div[2]""").get_attribute("innerHTML")
        except NoSuchElementException:
            pass
        else:
            if Ri == "No related items found.":
                writeData(path, "Sheet1", r, 3 , "Bin Ticket not added")
                continue
            else:
                pass
    elif priv is True:
        writeData(path, "Sheet1", r, 3 , "Vendor Connect")
        continue
    else:
        writeData(path, "Sheet1", r, 3 , "Not Bin related")
        continue
    try:
        B_Num = driver.find_element(By.XPATH, """/html/body/div[1]/form/div[2]/div/div[6]/div[2]/table/tbody/tr/td[3]/a""").click()
    except NoSuchElementException:
        writeData(path, "Sheet1", r, 3 , "Bin TT Not attached")
        continue
    else:
        pass
    time.sleep(4)
    driver.find_element(By.ID, "action_bar").click()
    time.sleep(5)
    try:
        status = driver.find_element(By.XPATH, """//*[@id="left_nav"]/tbody/tr[1]/td/div/span""").get_attribute("innerHTML")
    except NoSuchElementException:
        pass
    else:
        pass
    if status == "Resolved" or "Closed":
        pass
    else:
        Reminder = "Request you to complete the bin check and provide an update"
        driver.find_element(By.ID, "tab_correspondence").click()
        time.sleep(2)
        driver.find_element(By.ID, "correspondence").send_keys("Hi Team," + "\n\n" + Reminder + "\n\n" + "Thanks")
        # driver.find_element(By.ID, "save-button").click()
        writeData(path, "Sheet1", r, 3 , "Not resolved Reminder sent")
        writeData(path, "Sheet1", r, 1, Ticket_num)
        continue
    try:
        Root_cause = driver.find_element(By.XPATH, """//*[@id="root_cause_row"]/td/div/span""").get_attribute("innerHTML")
        print(Root_cause)
    except NoSuchElementException:
        pass
    else:
        pass
    try:
        driver.find_element(By.ID, "tab_correspondence").click()
    except NoSuchElementException:
        continue
    else:
        pass
    try:
        Contents = driver.find_element(By.CLASS_NAME, "text_to_format").text
    except NoSuchElementException:
        continue
    else:
        pass
    Content = Contents.lower()
    Msg = "\n No Defect Found"
    Msg2 = "\n Defect Found"
    Msg3 = "\n Other Detail"
    Msg4 = "\n NO Inventory for Bin Check"

    if Root_cause == "Defect Not Found - Information Provided" or "Defect Not Found - No Action Taken":
        writeData(path, "Sheet1", r, 3 , "No defect")
        writeData(path, "Sheet1", r, 4 , "---Bin-Check-Followup---" + "\n\n" + Msg + "\n\n" + "Bin-Check-Message" + "\n" + Content)
        txt = driver.find_element(By.ID, "u_text_search").send_keys(Ticket_num)
        driver.find_element(By.CSS_SELECTOR, "#u_remedy_lookup > button").click()
        time.sleep(4)
        driver.find_element(By.LINK_TEXT, "WW-Concession-Reduction").click()
        time.sleep(4)
        driver.find_element(By.ID, "tab_work_log").click()
        time.sleep(2)
        driver.find_element(By.ID, "work_log").send_keys("---Bin-Check-Followup---"+ "\n\n" + Msg + "\n\n" + "Bin-Check-Message" + "\n" + Content)
        continue

    elif Root_cause == "No Bin Checked - 50 Unit Limit" or "No Bin Checked - ASIN Flip Policy" or "No Bin Checked - Customer Damage" or "No Bin checked - Customer Order Attached" or "No Bin Checked - Duplicate Ticket" or "No Bin Checked - Hazmat UTC" or "No Bin Checked - Hazmat UTC" or "No Bin Checked – Incorrect CTI" or "No Bin Checked – ISS Cannot Perform" or "No Bin Checked - No Inventory" or "No Bin Checked - Normal FC Processing" or "No Bin Checked - Quarantined" or "No Bin Checked - Request Unclear" or "No Bin Checked - Transship" or "No Bin Checked – Under Review by Another Team" or "No Bin Checked - Unsafe to Complete" :
        writeData(path, "Sheet1", r, 3, "Check")
        writeData(path, "Sheet1", r, 4,"---Bin-Check-Followup---" + "\n\n" + Msg4 + "\n\n" + "Bin-Check-Message" + "\n" + Content)
        txt = driver.find_element(By.ID, "u_text_search").send_keys(Ticket_num)
        driver.find_element(By.CSS_SELECTOR, "#u_remedy_lookup > button").click()
        time.sleep(2)
        driver.find_element(By.LINK_TEXT, "WW-Concession-Reduction").click()
        time.sleep(4)
        driver.find_element(By.ID, "tab_work_log").click()
        time.sleep(2)
        driver.find_element(By.ID, "work_log").send_keys("---Bin-Check-Followup---" + "\n\n" + Msg4 + "\n\n" + "Bin-Check-Message" + "\n" + Content)
        continue

    elif Root_cause == "Other - See Details":
        writeData(path, "Sheet1", r, 3 , "Check")
        writeData(path, "Sheet1", r, 4 ,"---Bin-Check-Followup---" + "\n\n" + Msg3 + "\n\n" + "Bin-Check-Message" + "\n" + Content)
        txt = driver.find_element(By.ID, "u_text_search").send_keys(Ticket_num)
        driver.find_element(By.CSS_SELECTOR, "#u_remedy_lookup > button").click()
        time.sleep(2)
        driver.find_element(By.LINK_TEXT, "WW-Concession-Reduction").click()
        time.sleep(4)
        driver.find_element(By.ID, "tab_work_log").click()
        time.sleep(2)
        driver.find_element(By.ID, "work_log").send_keys("---Bin-Check-Followup---" + "\n\n" + Msg3 + "\n\n" + "Bin-Check-Message" + "\n" + Content)
        continue

    elif Root_cause == "Defect Confirmed - Adjustment Complete" or "Defect Confirmed - Box Change And Prep Required" or "Defect Confirmed - Box Change Required" or "Defect Confirmed - Broken Set" or "Defect Confirmed - Catalog Defect" or "Defect Confirmed - Information Provided" or "Defect Confirmed - Master Pack" or "Defect Confirmed - No Action Taken" or "Defect Confirmed - Other Change Required" or "Defect Confirmed - Prep Required" or "Defect Confirmed - Remove SIOC" or "Defect Confirmed - Remove SIOC and Prep Required" or "Defect Confirmed - Vendor Packaging Change" or "Defect Not Found - Information Provided":
        writeData(path, "Sheet1", r, 2, "Defect Confirmed")
        writeData(path, "Sheet1", r, 3,"---Bin-Check-Followup---" + "\n\n" + Msg2 + "\n\n" + "Bin-Check-Message" + "\n" + Content)
        txt = driver.find_element(By.ID, "u_text_search").send_keys(Ticket_num)
        driver.find_element(By.CSS_SELECTOR, "#u_remedy_lookup > button").click()
        time.sleep(4)
        driver.find_element(By.LINK_TEXT, "WW-Concession-Reduction").click()
        time.sleep(4)
        alaska = driver.find_element(By.PARTIAL_LINK_TEXT, "https://alaska-")
        act = ActionChains(driver)
        act.key_down(Keys.CONTROL).click(alaska).perform()
        driver.find_element(By.ID, "tab_work_log").click()
        time.sleep(3)
        Bin_creator = driver.find_element(By.PARTIAL_LINK_TEXT, "https://tt.amazon.com/new?")
        act.key_down(Keys.CONTROL).click(Bin_creator).perform()
        time.sleep(5)
        driver.switch_to.window(driver.window_handles[-1])
        time.sleep(3)
        try:
            fc = driver.find_element(By.XPATH, """//*[@id="view"]/table[2]""").get_attribute("outerHTML")
        except:
            writeData(path, "Sheet1", r, 4, "No Inventory")
            driver.switch_to.window(driver.window_handles[0])
            time.sleep(3)
            driver.find_element(By.ID, "tab_related_items").click()
            time.sleep(4)
            # tagging related bin item
            driver.find_element(By.ID, "tab_work_log").click()
            time.sleep(4)
            driver.find_element(By.ID, "work_log").send_keys("---Bin-Check-Followup---" + "\n\n" + Msg2 + "\n\n" + "NO inventory for Bin check")
            time.sleep(5)
            continue
        Fc_table = pd.read_html(fc)[0]
        FC = Fc_table[["FC", "Supply"]]
        FC.to_excel(r'C:\Users\kamesk\Desktop\Bin\Bin_check_File_FC.xlsx', index=1, header=None)
        driver.switch_to.window(driver.window_handles[1])
        time.sleep(3)
        driver.refresh()
        time.sleep(3)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(3)

        C_url = driver.current_url

        for P2 in range(2, rows2 + 1):
            try:
                writeData(path2, "Sheet1", P2, 399, " ")
            except PermissionError:
                driver.execute_script("alert('PLEASE CLOSE THE BIN_FC EXCEL AND RUN AGAIN');")
                time.sleep(3)
            else:
                pass
            Build_Id = readData(path2, "Sheet1", P2, 2)
            Fc_num = readData(path2, "Sheet1", P2, 3)
            Related_id = readData(path2,"Sheet1", P2, 11)
            if Build_Id == None:
                driver.execute_script("alert('Process Completed');")
                break
            elif Build_Id == "Total":
                driver.switch_to.window(driver.window_handles[0])
                time.sleep(3)
                driver.find_element(By.ID, "tab_related_items").click()
                time.sleep(4)
                #tagging related bin item

                driver.find_element(By.ID, "tab_work_log").click()
                time.sleep(4)
                driver.find_element(By.ID, "work_log").send_keys("---Bin-Check-Followup---" + "\n\n" + Msg2 + "\n\n" + "Bin-Check-Message" + "\n" + Content)
                time.sleep(5)
                continue
            else:
                pass
            if Fc_num > 5:
                pass
            else:
                continue
            driver.find_element(By.ID, "building_id").clear()
            time.sleep(3)
            try:
                Buil = driver.find_element(By.ID, "building_id").send_keys(Build_Id)
            except:
                Buil = driver.find_element(By.ID, "building_id").send_keys(Build_Id)
            else:
                pass
            time.sleep(2)
            driver.find_element(By.ID, "details").click()
            time.sleep(3)
            driver.find_element(By.ID, "tab-vendor-link").click()
            time.sleep(2)
            driver.find_element(By.ID, "quantity").clear()
            time.sleep(2)
            driver.find_element(By.ID, "quantity").send_keys(Fc_num)
            time.sleep(2)
            # driver.find_element(By.LINK_TEXT, "Submit Ticket").click()
            # time.sleep(10)
            # Bin_ID = driver.find_element(By.XPATH, """/html/body/div[2]/div[4]/table/tbody/tr/td[1]""").get_attribute("innerHTML")
            # writeData(path2, "Sheet1", P2, 11, Bin_ID)
            driver.get(C_url)
            time.sleep(3)

    else:
        writeData(path, "Sheet1", r, 3 , "Check required")
        writeData(path, "Sheet1", r, 4 ,"---Bin-Check-Followup---" + "\n\n" + Msg2 + "\n\n" + "Bin-Check-Message" + "\n" + Content)
        pass
        txt = driver.find_element(By.ID, "u_text_search").send_keys(Ticket_num)
        driver.find_element(By.CSS_SELECTOR, "#u_remedy_lookup > button").click()
        time.sleep(4)
        driver.find_element(By.LINK_TEXT, "WW-Concession-Reduction").click()
        time.sleep(4)
        driver.find_element(By.ID, "tab_work_log").click()
        time.sleep(2)
        driver.find_element(By.ID, "work_log").send_keys("---Bin-Check-Followup---" +"\n\n" + Msg2 + "\n\n" + "Bin-Check Message" + "\n" + Content)
        continue



